""" -----------------------------------------------------
# Copyright (c) [2022] Nadege LEMPERIERE
# All rights reserved
# -------------------------------------------------------
# Simulation scenario management
# -------------------------------------------------------
# Nadège LEMPERIERE, @04 november 2022
# Latest revision: 04 november 2022
# --------------------------------------------------- """

# Openpyxl includes
from openpyxl import load_workbook

class Context() :
    """ Singleton class managing a simulation context """

    m_data              = {}
    m_commands          = {}
    m_current_step      = -1

    def __new__(cls):

        if not hasattr(cls, 'instance'):
            cls.instance = super(Context, cls).__new__(cls)
        return cls.instance

    def load(self, filename, sheet) :
        """ Read scenario from excel file
        ---
        filename (str)  : Excel file in which the scenario data are located
        sheet (str)     : Sheet from which data shall be retrieved
        """

        self.m_data = {}

        # Load workbook with value rather than formula
        wbook = load_workbook(filename, data_only = True)

        # Select sheet
        content_sheet = wbook[sheet]

        # Associate header to column
        i_column = 1
        column_to_header = {}
        header_to_column = {}
        content = content_sheet.cell(1,i_column).value
        while content is not None :
            column_to_header[i_column]  = content
            header_to_column[content]   = i_column
            self.m_data[content] = []
            i_column += 1
            content = content_sheet.cell(1,i_column).value
        if 'time' not in header_to_column : raise Exception('Time column not found')

        for i_row in range(2,content_sheet.max_row + 1) :
            for col,header in column_to_header.items() :
                value = content_sheet.cell(i_row,col).value
                if isinstance(value,str) and value == 'True'  : value = True
                if isinstance(value,str) and value == 'False' : value = False
                self.m_data[header].append(value)

        self.restart()

    def step(self) :
        """ Take one step in simulation """
        self.m_current_step += 1
        self.m_commands['time'].append(self.get_data('time'))

    def restart(self) :
        """ Restart simulation """
        self.m_current_step = -1
        self.m_commands = {}
        self.m_commands['time']     = []
        self.m_commands['commands'] = []

    def log_command(self, command, args) :
        """ Push another command in
        ---
        command (str)   : Command to log
        args (dict)     : args to log
        """

        for _ in range(len(self.m_commands['time']) - len(self.m_commands['commands']) - 1) :
            self.m_commands['commands'].append(None)
        self.m_commands['commands'].append(command)

        for name, value in args.items() :
            if not name in self.m_commands :
                self.m_commands[name] = []
            for _ in range(len(self.m_commands['time']) - len(self.m_commands[name]) - 1) :
                self.m_commands[name].append(value)

    def get_data(self, column) :
        """ Get current data for current column
        ---
        column (str)  : Column to get data from
        ---
        returns       : Column data in its associated type
        """
        if column not in self.m_data :
            raise ValueError('Column ' + column + ' not found in data')
        return self.m_data[column][self.m_current_step]

    def get_commands(self) :
        """ Get logged commands
        ---
        returns (dict)      : All logged commands
        """
        return self.m_commands
