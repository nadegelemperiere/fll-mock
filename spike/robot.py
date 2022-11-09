""" -----------------------------------------------------
# Copyright (c) [2022] Nadege LEMPERIERE
# All rights reserved
# -------------------------------------------------------
# Simulation scenario management
# -------------------------------------------------------
# Nadège LEMPERIERE, @04 november 2022
# Latest revision: 04 november 2022
# --------------------------------------------------- """

# system includes
from json import load

# Openpyxl includes
from openpyxl import load_workbook

# Local includes
from spike.mock import Mock

class Robot(Mock) :
    """ Singleton class managing a simulation context """

    m_data = {}
    m_configuration = {}
    m_components = {}

    m_x_position = 0
    m_y_position = 0
    m_z_position = 0

    def __new__(cls):

        if not hasattr(cls, 'instance'):
            cls.instance = super(Robot, cls).__new__(cls)
        return cls.instance

    def __init__(self) :
        """ Contructor """
        self.s_default_columns  = {
            'x':'x',
            'y':'y',
            'z':'z',
        }
        if len(self.m_configuration) == 0 :
            super().__init__()


    def load_configuration(self, configuration) :
        """ Loading robot static configuration
        ---
        configuration (str) : Json file containing robots parameters
        """
        with open(configuration,'r', encoding='UTF-8') as file :
            self.m_configuration = load(file)
            file.close()
        print(str(self.m_configuration))

    def load_scenario(self, filename, sheet) :
        """ Read scenario from excel file
        ---
        filename (str)  : Excel file in which the scenario data are located
        sheet (str)     : Sheet from which data shall be retrieved
        """

        self.m_data = {}

        # Load workbook with value rather than formula
        wbook = load_workbook(filename, data_only = True)

        # Select sheet
        content_sheet = wbook[sheet]

        # Associate header to column
        i_column = 1
        column_to_header = {}
        header_to_column = {}
        content = content_sheet.cell(1,i_column).value
        while content is not None :
            column_to_header[i_column]  = content
            header_to_column[content]   = i_column
            self.m_data[content] = []
            i_column += 1
            content = content_sheet.cell(1,i_column).value
        if 'time' not in header_to_column : raise Exception('Time column not found')

        for i_row in range(2,content_sheet.max_row + 1) :
            for col,header in column_to_header.items() :
                value = content_sheet.cell(i_row,col).value
                self.m_data[header].append(value)

    def step(self) :

        self.m_x_position = int(self.m_scenario['x'][self.m_current_step])
        self.m_y_position = int(self.m_scenario['y'][self.m_current_step])
        self.m_z_position = int(self.m_scenario['z'][self.m_current_step])

        super().step()

    def initialize(self, data) :
        """ Loading a mock with scenario simulated data
        ---
        data (obj)  : Mock object to initialize
        """
        data.initialize(self.m_data)

    def register_component(self, port, component) :
        """ Return type of component hosted by the port
        ---
        port (str)      : port to check
        component (obj) : the component hosted
        """

        if port in self.m_configuration['ports'] :
            test_type = "<class 'spike." + self.m_configuration['ports'][port].lower() + \
                '.' + self.m_configuration['ports'][port] + "'>"
            if str(type(component)) == test_type :
                self.m_components[port] = component
            else :
                raise TypeError('component does not fit in port ' + port)
        else:
            raise TypeError('component does not fit in port ' + port)

    def get_component(self, port) :
        """ Return type of component hosted by the port
        ---
        port (str)      : port to check
        ---
        returns (str)   : the type of component hosted
        """

        result = None
        if port in self.m_configuration['ports'] :
            result = self.m_configuration['ports'][port]
        return result
